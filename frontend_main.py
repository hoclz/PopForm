import sys
# Force Python to see E:\AI\app so that all modules are recognized.
sys.path.append(r"E:\AI\app")

from tkinter import ttk, messagebox
import tkinter as tk
import os
import webbrowser
import pandas as pd
import tempfile

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Import modules directly (all files are in the same folder)
import backend_main_processing
import frontend_data_loader
import frontend_bracket_utils
import frontend_report_window

# Path constants
DATA_FOLDER = r"E:\AI\app\data"
FORM_CONTROL_PATH = r"E:\AI\app\form_control_UI_data.csv"

# Global DataFrame for "Download Output"
filtered_data_for_download = pd.DataFrame()

# Global list of results for multi-year display (needed for multi-sheet Excel)
results_for_display_global = []

# --- Race Mapping ---
# Mapping from friendly display names to race codes
RACE_DISPLAY_TO_CODE = {
    "Two or More Races": "TOM",
    "American Indian and Alaska Native": "AIAN",
    "Black or African American": "Black",
    "White": "White",
    "Native Hawaiian and Other Pacific Islander": "NHOPI",
    "Asian": "Asian"
}

# --- AgeGroup Mapping ---
# We'll show friendlier names in the UI but still pass original codes to backend.
AGEGROUP_DISPLAY_TO_CODE = {
    "All": "All",          # Keep "All" as is
    "18-Bracket": "agegroup13",
    "6-Bracket":  "agegroup14",
    "2-Bracket":  "agegroup15"
}

################################################################################
# 1) A small helper dict so we can map integer codes (1..18) back to bracket labels.
#    This matches your BRACKET_MAP in frontend_bracket_utils.py but inverted.
CODE_TO_BRACKET = {
    1:  "0-4",
    2:  "5-9",
    3:  "10-14",
    4:  "15-19",
    5:  "20-24",
    6:  "25-29",
    7:  "30-34",
    8:  "35-39",
    9:  "40-44",
    10: "45-49",
    11: "50-54",
    12: "55-59",
    13: "60-64",
    14: "65-69",
    15: "70-74",
    16: "75-79",
    17: "80-84",
    18: "80+"
}

def combine_codes_to_label(codes: list[int]) -> str:
    """
    Given a list of integer codes, e.g. [1,2,3,4],
    produce a single bracket label by merging the lower and upper ends
    of each code's bracket. For example, [1..4] => "0-19".
    """
    # Sort & unique
    codes = sorted(set(codes))
    if not codes:
        return ""

    # For each code, parse the bracket label from CODE_TO_BRACKET
    # e.g. code=1 => "0-4", code=2 => "5-9", etc.
    # We'll find the min of all lower bounds and max of all upper bounds
    # to produce a single combined bracket like "0-19".
    low_vals = []
    high_vals = []
    for c in codes:
        bracket_str = CODE_TO_BRACKET.get(c, "")
        if "-" in bracket_str:
            parts = bracket_str.split("-")
            try:
                start = int(parts[0])
                # check if it ends with "+" e.g. "80+"
                if parts[1].endswith("+"):
                    # e.g. "80+"
                    end = int(parts[1].replace("+", ""))
                    # We'll treat that as effectively 999 for combining
                    end = 999
                else:
                    end = int(parts[1])
                low_vals.append(start)
                high_vals.append(end)
            except:
                pass
        elif bracket_str.endswith("+"):
            # e.g. "80+"
            try:
                start = int(bracket_str.replace("+", ""))
                low_vals.append(start)
                high_vals.append(999)
            except:
                pass

    if not low_vals or not high_vals:
        # fallback: just show code numbers
        return "-".join(str(c) for c in codes)

    overall_low = min(low_vals)
    overall_high = max(high_vals)

    if overall_high >= 999:
        return f"{overall_low}+"
    else:
        return f"{overall_low}-{overall_high}"
################################################################################

def main():
    global filtered_data_for_download, results_for_display_global

    # 1) Load form-control data
    (years_list,
     agegroups_list_raw,
     races_list_raw,
     counties_map,
     agegroup_map_explicit,
     agegroup_map_implicit) = frontend_data_loader.load_form_control_data(FORM_CONTROL_PATH)

    # Build a user-friendly AgeGroup list
    agegroup_display_list = list(AGEGROUP_DISPLAY_TO_CODE.keys())  # ["All", "18-Bracket", "6-Bracket", "2-Bracket"]

    # Build a user-friendly Race list
    race_display_list = ["All"]  # always include "All" at the top
    for rcode in sorted(races_list_raw):
        if rcode == "All":
            continue
        friendly_name = None
        for k, v in RACE_DISPLAY_TO_CODE.items():
            if v == rcode:
                friendly_name = k
                break
        race_display_list.append(friendly_name if friendly_name else rcode)

    root = tk.Tk()
    root.title("Illinois Census Data Form")
    root.geometry("1000x600")

    # UI Layout
    title_label = tk.Label(root, text="Illinois Census Data Form", font=("Arial", 16, "bold"))
    title_label.grid(row=0, column=0, columnspan=4, padx=10, pady=10, sticky="w")

    main_frame = tk.Frame(root)
    main_frame.grid(row=1, column=0, columnspan=4, sticky="nsew")
    root.grid_rowconfigure(1, weight=1)
    for c in range(4):
        root.grid_columnconfigure(c, weight=1)

    # LEFT PANEL - Years & Counties
    left_panel = tk.Frame(main_frame, bd=2, relief="groove", padx=10, pady=10)
    left_panel.grid(row=0, column=0, sticky="nsew")
    main_frame.grid_columnconfigure(0, weight=1)

    tk.Label(left_panel, text="Year Selection:").pack(anchor="w")
    year_listbox_frame = tk.Frame(left_panel)
    year_listbox_frame.pack(fill="both", expand=True)
    listbox_years = tk.Listbox(year_listbox_frame, selectmode="extended", height=8, exportselection=0)
    listbox_years.pack(side="left", fill="both", expand=True)
    scrollbar_years = tk.Scrollbar(year_listbox_frame, orient="vertical", command=listbox_years.yview)
    scrollbar_years.pack(side="right", fill="y")
    listbox_years.config(yscrollcommand=scrollbar_years.set)

    for yr in years_list:
        listbox_years.insert(tk.END, yr)
    if years_list:
        listbox_years.selection_set(0)

    tk.Label(left_panel, text="Select Counties:").pack(anchor="w")
    county_listbox_frame = tk.Frame(left_panel)
    county_listbox_frame.pack(fill="both", expand=True)
    listbox_counties = tk.Listbox(county_listbox_frame, selectmode="extended", height=10, exportselection=0)
    listbox_counties.pack(side="left", fill="both", expand=True)
    scrollbar_counties = tk.Scrollbar(county_listbox_frame, orient="vertical", command=listbox_counties.yview)
    scrollbar_counties.pack(side="right", fill="y")
    listbox_counties.config(yscrollcommand=scrollbar_counties.set)
    listbox_counties.insert(tk.END, "All")
    for cname in sorted(counties_map.keys()):
        listbox_counties.insert(tk.END, cname)

    # MIDDLE PANEL - AgeGroup, Race, Ethnicity, Sex, Region
    middle_panel = tk.Frame(main_frame, bd=2, relief="groove", padx=10, pady=10)
    middle_panel.grid(row=0, column=1, sticky="nsew")
    main_frame.grid_columnconfigure(1, weight=1)

    tk.Label(middle_panel, text="Age Group:").pack(anchor="w")
    combo_agegroup = ttk.Combobox(middle_panel, values=agegroup_display_list, state="readonly")
    combo_agegroup.pack(fill="x", pady=5)
    combo_agegroup.set("All")

    tk.Label(middle_panel, text="Selected Age Brackets (Implicit):").pack(anchor="w")
    bracket_text_frame = tk.Frame(middle_panel)
    bracket_text_frame.pack(fill="both", expand=True)
    text_brackets = tk.Text(bracket_text_frame, height=8, wrap="word", state="disabled")
    text_brackets.pack(side="left", fill="both", expand=True)
    scrollbar_brackets = tk.Scrollbar(bracket_text_frame, orient="vertical", command=text_brackets.yview)
    scrollbar_brackets.pack(side="right", fill="y")
    text_brackets.config(yscrollcommand=scrollbar_brackets.set)

    tk.Label(middle_panel, text="Race:").pack(anchor="w")
    combo_race = ttk.Combobox(middle_panel, values=race_display_list, state="readonly")
    combo_race.pack(fill="x", pady=5)
    combo_race.current(0)

    eth_frame = tk.LabelFrame(middle_panel, text="Ethnicity", padx=5, pady=5)
    eth_frame.pack(fill="x", pady=5)
    ethnicity_var = tk.StringVar(value="All")
    tk.Radiobutton(eth_frame, text="All", variable=ethnicity_var, value="All").pack(anchor="w")
    tk.Radiobutton(eth_frame, text="Hispanic", variable=ethnicity_var, value="Hispanic").pack(anchor="w")
    tk.Radiobutton(eth_frame, text="Not Hispanic", variable=ethnicity_var, value="Not Hispanic").pack(anchor="w")

    sex_frame = tk.LabelFrame(middle_panel, text="Sex", padx=5, pady=5)
    sex_frame.pack(fill="x", pady=5)
    sex_var = tk.StringVar(value="All")
    tk.Radiobutton(sex_frame, text="All", variable=sex_var, value="All").pack(anchor="w")
    tk.Radiobutton(sex_frame, text="Male", variable=sex_var, value="Male").pack(anchor="w")
    tk.Radiobutton(sex_frame, text="Female", variable=sex_var, value="Female").pack(anchor="w")

    region_frame = tk.LabelFrame(middle_panel, text="Regional Counties", padx=5, pady=5)
    region_frame.pack(fill="x", pady=5)
    region_var = tk.StringVar(value="None")
    tk.Radiobutton(region_frame, text="None", variable=region_var, value="None").pack(anchor="w")
    tk.Radiobutton(region_frame, text="Collar", variable=region_var, value="Collar Counties").pack(anchor="w")
    tk.Radiobutton(region_frame, text="Urban", variable=region_var, value="Urban Counties").pack(anchor="w")
    tk.Radiobutton(region_frame, text="Rural", variable=region_var, value="Rural Counties").pack(anchor="w")

    # RIGHT PANEL - Custom Age Ranges
    right_panel = tk.Frame(main_frame, bd=2, relief="groove", padx=10, pady=10)
    right_panel.grid(row=0, column=2, sticky="nsew")
    main_frame.grid_columnconfigure(2, weight=1)

    age_range_frame = tk.LabelFrame(right_panel, text="Custom Age Ranges (Min / Max)", padx=5, pady=5)
    age_range_frame.pack(fill="x", pady=5)
    custom_age_entries = []
    for i in range(1, 6):
        row_frame = tk.Frame(age_range_frame)
        row_frame.pack(fill="x", pady=2)
        lbl_min = tk.Label(row_frame, text=f"Min{i}:")
        lbl_min.pack(side="left")
        entry_min = tk.Entry(row_frame, width=5)
        entry_min.pack(side="left", padx=2)
        lbl_max = tk.Label(row_frame, text=f"Max{i}:")
        lbl_max.pack(side="left")
        entry_max = tk.Entry(row_frame, width=5)
        entry_max.pack(side="left", padx=2)
        custom_age_entries.append((entry_min, entry_max))

    # Add a note about custom ranges overriding Age Group
    lbl_override = tk.Label(
        age_range_frame,
        text="Note: Custom Age Ranges override the Age Group selection above.\nValid Age codes: 1..18"
    )
    lbl_override.pack(anchor="w", pady=5)

    bottom_frame = tk.Frame(root, padx=10, pady=10)
    bottom_frame.grid(row=2, column=0, columnspan=4, sticky="se")

    def on_agegroup_selected(event=None):
        selected_display = combo_agegroup.get().strip()
        text_brackets.config(state="normal")
        text_brackets.delete("1.0", tk.END)

        if not selected_display or selected_display == "All":
            text_brackets.insert(tk.END, "No Age Group selected.\n")
            text_brackets.config(state="disabled")
            return

        # Convert display name to code
        agegroup_code = AGEGROUP_DISPLAY_TO_CODE.get(selected_display, "All")

        brackets_implicit = agegroup_map_implicit.get(agegroup_code, [])
        if not brackets_implicit:
            text_brackets.insert(tk.END, f"No bracket expressions found for {selected_display}.\n")
        else:
            for expr in brackets_implicit:
                text_brackets.insert(tk.END, f"- {expr}\n")
        text_brackets.config(state="disabled")

    combo_agegroup.bind("<<ComboboxSelected>>", on_agegroup_selected)

    def on_generate_report():
        global filtered_data_for_download, results_for_display_global
        filtered_data_for_download = pd.DataFrame()
        results_for_display_global = []

        selected_years = [listbox_years.get(i) for i in listbox_years.curselection()]
        if not selected_years:
            messagebox.showwarning("Warning", "Please select at least one year.")
            return

        selected_counties_list = [listbox_counties.get(i) for i in listbox_counties.curselection()]

        # Convert Race display back to code, but keep the display name for final reporting
        selected_race_display = combo_race.get()
        if selected_race_display == "All":
            selected_race_code = "All"
        else:
            # Find the code from the mapping
            for k, v in RACE_DISPLAY_TO_CODE.items():
                if k == selected_race_display:
                    selected_race_code = v
                    break
            else:
                selected_race_code = selected_race_display  # fallback if not found

        selected_ethnicity = ethnicity_var.get()
        selected_sex = sex_var.get()
        selected_region = region_var.get()

        # Convert AgeGroup display back to code
        selected_agegroup_display = combo_agegroup.get()
        selected_agegroup_code = AGEGROUP_DISPLAY_TO_CODE.get(selected_agegroup_display, None)
        # If "All" is selected for Age Group, pass None to backend
        if selected_agegroup_code == "All":
            agegroup_for_backend = None
        else:
            agegroup_for_backend = selected_agegroup_code

        counties_str = ", ".join(selected_counties_list) if selected_counties_list else "No counties selected."

        # Validate custom ranges
        custom_ranges = []
        for (entry_min, entry_max) in custom_age_entries:
            min_val = entry_min.get().strip()
            max_val = entry_max.get().strip()
            if min_val.isdigit() and max_val.isdigit():
                mn = int(min_val)
                mx = int(max_val)
                # Enforce 1..18
                if not (1 <= mn <= 18 and 1 <= mx <= 18):
                    messagebox.showwarning(
                        "Invalid Range",
                        "Custom Age Ranges must be between 1 and 18 inclusive.\nNo data will be returned."
                    )
                    return  # Stop generating the report
                if mn <= mx:
                    custom_ranges.append((mn, mx))

        custom_age_str = "; ".join(f"{mn}-{mx}" for (mn, mx) in custom_ranges) if custom_ranges else "No custom age"

        combined_frames = []

        # For the final report's title, decide single or multiple years
        if len(selected_years) == 1:
            years_title = f"({selected_years[0]})"
        else:
            # e.g. (2000, 2001) or (Multiple Years)
            years_title = f"({', '.join(selected_years)})"

        for year_str in selected_years:
            df = backend_main_processing.process_population_data(
                data_folder=DATA_FOLDER,
                agegroup_map_explicit=agegroup_map_explicit,
                counties_map=counties_map,
                selected_years=[year_str],
                selected_counties=selected_counties_list,
                selected_race=selected_race_code,
                selected_ethnicity=selected_ethnicity,
                selected_sex=selected_sex,
                selected_region=selected_region,
                selected_agegroup=agegroup_for_backend,
                custom_age_ranges=custom_ranges
            )

            # If "All" => single row "IL Population" if no custom ranges
            if agegroup_for_backend is None and not custom_ranges:
                total_pop = df["Count"].sum() if not df.empty else 0
                out_df = pd.DataFrame({
                    "AgeGroup": ["IL Population"],
                    "Count": [total_pop],
                    "Percent": [100.0 if total_pop > 0 else 0.0],
                    "Year": [year_str]
                })
                results_for_display_global.append(
                    (
                        year_str,
                        out_df,
                        selected_agegroup_display,
                        selected_ethnicity,
                        selected_sex,
                        selected_region,
                        selected_race_display,  # pass the friendly race name
                        counties_str,
                        custom_age_str,
                        years_title
                    )
                )
                combined_frames.append(out_df)
            else:
                # If custom_ranges is non-empty, skip implicit bracket logic entirely
                if custom_ranges:
                    rows = []
                    total_sum = 0
                    for (mn, mx) in custom_ranges:
                        # We'll combine codes mn..mx into a single bracket label
                        code_list = range(mn, mx + 1)
                        bracket_label = combine_codes_to_label(code_list)
                        # Sum population for these codes
                        mask = df["Age"].isin(code_list)
                        sub_sum = df.loc[mask, "Count"].sum()
                        rows.append((bracket_label, sub_sum))
                        total_sum += sub_sum

                    out_rows = []
                    for (bexpr, cval) in rows:
                        pct = (cval / total_sum * 100.0) if total_sum > 0 else 0.0
                        out_rows.append((bexpr, cval, pct))

                    out_df = pd.DataFrame(out_rows, columns=["AgeGroup", "Count", "Percent"])
                    out_df["Percent"] = out_df["Percent"].round(1)
                    out_df["Year"] = year_str

                    combined_frames.append(out_df)
                    results_for_display_global.append(
                        (
                            year_str,
                            out_df,
                            selected_agegroup_display,
                            selected_ethnicity,
                            selected_sex,
                            selected_region,
                            selected_race_display,
                            counties_str,
                            custom_age_str,
                            years_title
                        )
                    )
                else:
                    # Use the implicit bracket expressions
                    brackets_implicit = agegroup_map_implicit.get(agegroup_for_backend, [])
                    if not brackets_implicit:
                        total_pop = df["Count"].sum() if not df.empty else 0
                        out_df = pd.DataFrame({
                            "AgeGroup": [f"No bracket for {selected_agegroup_display}"],
                            "Count": [total_pop],
                            "Percent": [100.0 if total_pop > 0 else 0.0],
                            "Year": [year_str]
                        })
                        results_for_display_global.append(
                            (
                                year_str,
                                out_df,
                                selected_agegroup_display,
                                selected_ethnicity,
                                selected_sex,
                                selected_region,
                                selected_race_display,
                                counties_str,
                                custom_age_str,
                                years_title
                            )
                        )
                        combined_frames.append(out_df)
                    else:
                        if "Age" in df.columns and not df.empty:
                            rows = []
                            total_sum = 0
                            for bracket_expr in brackets_implicit:
                                bracket_expr = bracket_expr.strip()
                                mask = frontend_bracket_utils.parse_implicit_bracket(df, bracket_expr)
                                sub_sum = df.loc[mask, "Count"].sum()
                                rows.append((bracket_expr, sub_sum))
                                total_sum += sub_sum
                        else:
                            rows = [(br.strip(), 0) for br in brackets_implicit]
                            total_sum = 0

                        out_rows = []
                        for (bexpr, cval) in rows:
                            pct = (cval / total_sum * 100.0) if total_sum > 0 else 0.0
                            out_rows.append((bexpr, cval, pct))

                        out_df = pd.DataFrame(out_rows, columns=["AgeGroup", "Count", "Percent"])
                        out_df["Percent"] = out_df["Percent"].round(1)
                        out_df["Year"] = year_str
                        combined_frames.append(out_df)
                        results_for_display_global.append(
                            (
                                year_str,
                                out_df,
                                selected_agegroup_display,
                                selected_ethnicity,
                                selected_sex,
                                selected_region,
                                selected_race_display,
                                counties_str,
                                custom_age_str,
                                years_title
                            )
                        )

        if combined_frames:
            final_combined = pd.concat(combined_frames, ignore_index=True)
        else:
            final_combined = pd.DataFrame(columns=["AgeGroup", "Count", "Percent", "Year"])

        filtered_data_for_download = final_combined

        # Check if any of the result DataFrames are non-empty
        any_nonempty = any(not tup[1].empty for tup in results_for_display_global)
        if not any_nonempty:
            messagebox.showinfo("Result", "No data found for the selected filters.")
        else:
            # Show multi-year report in a new window
            # We'll pass the 'years_title' inside each tuple so the report window can use it.
            frontend_report_window.show_multi_year_report_in_new_window(root, results_for_display_global)

    def on_clear_selections():
        listbox_years.selection_clear(0, tk.END)
        if years_list:
            listbox_years.selection_set(0)
            listbox_years.see(0)
        listbox_counties.selection_clear(0, tk.END)
        combo_agegroup.set("All")
        combo_race.current(0)
        text_brackets.config(state="normal")
        text_brackets.delete("1.0", tk.END)
        text_brackets.insert(tk.END, "No Age Group selected.\n")
        text_brackets.config(state="disabled")
        ethnicity_var.set("All")
        sex_var.set("All")
        region_var.set("None")
        for (emin, emax) in custom_age_entries:
            emin.delete(0, tk.END)
            emax.delete(0, tk.END)

    def on_clear_report():
        messagebox.showinfo("Clear Report", "Report cleared (placeholder).")

    def on_close():
        if messagebox.askyesno("Confirm", "Are you sure you want to close?"):
            root.destroy()

    def on_download_output():
        global filtered_data_for_download, results_for_display_global
        if filtered_data_for_download.empty:
            messagebox.showinfo("Download", "No data to download. Please generate a report first.")
            return

        dl_win = tk.Toplevel(root)
        dl_win.title("Download Output")
        dl_win.geometry("300x200")
        label_dl = tk.Label(dl_win, text="Choose the output format for download:")
        label_dl.pack(pady=10)
        dl_var = tk.StringVar(value="CSV")
        tk.Radiobutton(dl_win, text="CSV", variable=dl_var, value="CSV").pack(anchor="w")
        tk.Radiobutton(dl_win, text="Excel", variable=dl_var, value="Excel").pack(anchor="w")

        def do_download():
            fmt = dl_var.get()
            if filtered_data_for_download.empty:
                messagebox.showinfo("Download", "No data to download.")
                dl_win.destroy()
                return

            if fmt == "CSV":
                # Single CSV with appended rows
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".csv")
                tmp_name = tmp.name
                tmp.close()
                filtered_data_for_download.to_csv(tmp_name, index=False)
                webbrowser.open(tmp_name)
                messagebox.showinfo("Download", f"CSV opened:\n{tmp_name}")
            elif fmt == "Excel":
                # Multi-sheet Excel to mirror multi-year display
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                tmp_name = tmp.name
                tmp.close()

                wb = openpyxl.Workbook()
                # Remove default sheet
                wb.remove(wb.active)

                # Each tuple in results_for_display_global is: 
                # (year_str, df_out, agegroup_display, ethnicity, sex, region, race_display, counties_str, custom_age_str, years_title)
                for (year_str, df_out, *_rest) in results_for_display_global:
                    if df_out.empty:
                        continue
                    # If sheet name already exists, add suffix
                    sheet_name = str(year_str)
                    count_dup = 2
                    while sheet_name in wb.sheetnames:
                        sheet_name = f"{year_str}_{count_dup}"
                        count_dup += 1

                    ws = wb.create_sheet(title=sheet_name)
                    # Write df_out to sheet
                    for r_idx, row in enumerate(dataframe_to_rows(df_out, index=False, header=True), 1):
                        for c_idx, val in enumerate(row, 1):
                            ws.cell(row=r_idx, column=c_idx, value=val)

                wb.save(tmp_name)
                webbrowser.open(tmp_name)
                messagebox.showinfo("Download", f"Excel opened:\n{tmp_name}")

            dl_win.destroy()

        btn_confirm = ttk.Button(dl_win, text="Download", command=do_download)
        btn_confirm.pack(pady=10)

    def on_show_census_links():
        top = tk.Toplevel(root)
        top.title("Census Data Links")
        top.geometry("700x300")
        label_info = tk.Label(top, text=(
            "Below are important links to access the data information.\n"
            "They will be updated from time to time if needed.\n"
            "Select one or more links and click 'Open Selected Link(s)'."
        ))
        label_info.pack(pady=5)
        links_frame = tk.Frame(top)
        links_frame.pack(fill="both", expand=True)
        listbox_links = tk.Listbox(links_frame, selectmode="extended", height=8)
        listbox_links.pack(side="left", fill="both", expand=True)
        scrollbar_links = tk.Scrollbar(links_frame, orient="vertical", command=listbox_links.yview)
        scrollbar_links.pack(side="right", fill="y")
        listbox_links.config(yscrollcommand=scrollbar_links.set)

        links_data = [
            "https://www2.census.gov/programs-surveys/popest/datasets/",
            "https://www2.census.gov/programs-surveys/popest/datasets/2000-2010/intercensal/county/",
            "https://www2.census.gov/programs-surveys/popest/datasets/2010-2020/counties/asrh/",
            "https://www2.census.gov/programs-surveys/popest/datasets/2020-2023/counties/asrh/",
            "https://www2.census.gov/programs-surveys/popest/datasets/2020-2024/counties/asrh/",
            "RELEASE SCHEDULE: https://www.census.gov/programs-surveys/popest/about/schedule.html"
        ]
        for link in links_data:
            listbox_links.insert(tk.END, link)

        def open_selected_links():
            selected_indices = listbox_links.curselection()
            for idx in selected_indices:
                link = listbox_links.get(idx)
                if link.startswith("RELEASE SCHEDULE: "):
                    link = link.replace("RELEASE SCHEDULE: ", "")
                webbrowser.open(link)

        btn_open = ttk.Button(top, text="Open Selected Link(s)", command=open_selected_links)
        btn_open.pack(pady=5)

    btn_generate = ttk.Button(bottom_frame, text="Generate Report", command=on_generate_report)
    btn_generate.pack(side="left", padx=5)

    btn_clear_sel = ttk.Button(bottom_frame, text="Clear Selections", command=on_clear_selections)
    btn_clear_sel.pack(side="left", padx=5)

    btn_clear_report = ttk.Button(bottom_frame, text="Clear Report", command=on_clear_report)
    btn_clear_report.pack(side="left", padx=5)

    btn_links = ttk.Button(bottom_frame, text="Census Links", command=on_show_census_links)
    btn_links.pack(side="left", padx=5)

    btn_download = ttk.Button(bottom_frame, text="Download Output", command=on_download_output)
    btn_download.pack(side="left", padx=5)

    btn_close = ttk.Button(bottom_frame, text="Close", command=on_close)
    btn_close.pack(side="left", padx=5)

    text_brackets.config(state="normal")
    text_brackets.delete("1.0", tk.END)
    text_brackets.insert(tk.END, "No Age Group selected.\n")
    text_brackets.config(state="disabled")

    root.mainloop()

if __name__ == "__main__":
    main()
